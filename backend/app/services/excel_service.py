import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.validation_utils import parse_number, clean_item_name

NAVY = "1B3E5F"
MID_BLUE = "2E75B6"
LT_BLUE = "D6E4F0"
SEC_BG = "E8F0F7"
PALE = "F2F2F2"
WHITE = "FFFFFF"
BLUE_NUM = "0000FF"
BLACK = "000000"
DARK_GREY = "444444"
GREEN = "008000"

_DOC_COLORS = {
    "pl": ("1B3E5F", "2E75B6"),
    "bs": ("1A4731", "2E7D52"),
    "cf": ("3D2B1F", "7B5430"),
    "insurance": ("4A235A", "7D3C98"),
    "generic": ("1B3E5F", "2E75B6"),
}

def _detect_doc_type(raw_rows) -> str:
    all_text = " ".join(r.get("item", "") for r in raw_rows[:30]).lower()
    if any(k in all_text for k in ["premium", "claims", "underwriting", "reinsurance"]):
        return "insurance"
    if any(k in all_text for k in ["total assets", "liabilities", "shareholders equity"]):
        return "bs"
    if any(k in all_text for k in ["operating activities", "investing activities", "financing activities"]):
        return "cf"
    if any(k in all_text for k in ["revenue from operations", "ebitda", "profit before tax"]):
        return "pl"
    return "generic"

def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def _border(color="CCCCCC"):
    s = _side(color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom(color="DDDDDD"):
    return Border(bottom=_side(color=color))

def _med_bottom(color=NAVY):
    return Border(bottom=_side(style="medium", color=color))

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _title_bar(ws, row, ncols, text, hdr_color=NAVY):
    ws.row_dimensions[row].height = 26
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, color=WHITE, name="Arial")
    c.fill = _fill(hdr_color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, ncols + 2):
        ws.cell(row=row, column=col).fill = _fill(hdr_color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols + 1)

def _header_row(ws, row, year_headers, hdr_color=MID_BLUE):
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=1, value="Particulars")
    c.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill = _fill(hdr_color)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _border(hdr_color)
    for i, h in enumerate(year_headers):
        cell = ws.cell(row=row, column=i + 2, value=h)
        cell.font = Font(bold=True, size=9, color=WHITE, name="Arial")
        cell.fill = _fill(hdr_color)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _border(hdr_color)

def _subheader_row(ws, row, n, unit_label="(₹ in crores)", hdr_color=MID_BLUE):
    ws.row_dimensions[row].height = 13
    ws.cell(row=row, column=1).fill = _fill(hdr_color)
    for i in range(n):
        c = ws.cell(row=row, column=i + 2, value=unit_label)
        c.font = Font(italic=True, size=8, color=WHITE, name="Arial")
        c.fill = _fill(hdr_color)
        c.alignment = Alignment(horizontal="right", vertical="center")

def _section_header(ws, row, ncols, text):
    ws.row_dimensions[row].height = 16
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=10, color=NAVY, name="Arial")
    c.fill = _fill(SEC_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(bottom=_side(color=MID_BLUE))
    for col in range(2, ncols + 2):
        cc = ws.cell(row=row, column=col)
        cc.fill = _fill(SEC_BG)
        cc.border = Border(bottom=_side(color=MID_BLUE))

def _data_row(ws, row, label, values, indent=0, bold=False):
    ws.row_dimensions[row].height = 15
    c = ws.cell(row=row, column=1, value=("  " * indent) + label)
    c.font = Font(bold=bold, size=9, color=BLACK if bold else DARK_GREY, name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _bottom()
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=i + 2, value=v)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _bottom()
        if isinstance(v, (int, float)):
            cell.font = Font(bold=bold, size=9, color=BLUE_NUM, name="Arial")
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        elif isinstance(v, str) and v.startswith("="):
            cell.font = Font(bold=bold, size=9, color=BLACK, name="Arial")
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        else:
            cell.font = Font(bold=bold, size=9, color=DARK_GREY, name="Arial")

def _total_row(ws, row, ncols, label, values, bg=PALE, double_border=False):
    ws.row_dimensions[row].height = 17
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, size=10, color=NAVY, name="Arial")
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _med_bottom() if double_border else _border(NAVY)
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=i + 2, value=v)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _med_bottom() if double_border else _border(NAVY)
        if isinstance(v, (int, float)):
            cell.font = Font(bold=True, size=10, color=NAVY, name="Arial")
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        elif isinstance(v, str) and v.startswith("="):
            cell.font = Font(bold=True, size=10, color=BLACK, name="Arial")
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
        else:
            cell.font = Font(bold=True, size=10, color=NAVY, name="Arial")

def _pct_row(ws, row, label, formulas):
    ws.row_dimensions[row].height = 13
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(italic=True, size=8, color="555555", name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border = _bottom("EEEEEE")
    for i, f in enumerate(formulas):
        cell = ws.cell(row=row, column=i + 2, value=f)
        cell.font = Font(italic=True, size=8, color=BLACK, name="Arial")
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = "0.0%"
        cell.border = _bottom("EEEEEE")

def _spacer(ws, row, height=5):
    ws.row_dimensions[row].height = height

_TOTAL_KW = [
    "total", "gross profit", "ebitda", "profit before", "profit after",
    "profit for", "net profit", "earnings", "pat", "pbt",
    "profit/(loss)", "loss for", "total comprehensive", "total income",
    "total revenue", "total expenses", "net premium", "net claims",
    "surplus", "deficit", "net cash",
]

_PAT_KW = [
    "profit after tax", "pat", "profit for the period",
    "profit for the year", "profit/(loss) for",
]

_SECTION_ONLY_KW = [
    "revenue from operations", "other income", "expenses",
    "cost of material", "purchases of stock", "employee benefit",
    "finance costs", "other expenses", "depreciation",
    "tax expense", "other comprehensive", "discontinued",
    "assets", "liabilities", "equity", "cash flow",
    "operating activities", "investing activities", "financing activities",
    "premium income", "claims incurred", "underwriting",
]

def _classify_row(item, has_values, is_section_header_flag):
    lower = item.lower().strip()
    if is_section_header_flag and not has_values:
        return "SECTION"
    if not has_values and any(lower == k or lower.startswith(k) for k in _SECTION_ONLY_KW):
        return "SECTION"
    if any(k in lower for k in _TOTAL_KW):
        return "TOTAL"
    return "DATA"

def _is_pat(item):
    lower = item.lower()
    return any(k in lower for k in _PAT_KW)

def _set_col_widths(ws, n):
    ws.column_dimensions["A"].width = 62
    for i in range(n):
        ws.column_dimensions[get_column_letter(i + 2)].width = 17

def _detect_doc_title(raw_rows, year_headers) -> str:
    all_text = " ".join(r.get("item", "") for r in raw_rows[:30]).lower()
    parts = []

    if "consolidated" in all_text:
        parts.append("Consolidated")
    elif "standalone" in all_text:
        parts.append("Standalone")

    if any(k in all_text for k in ["premium", "claims", "underwriting", "reinsurance"]):
        parts.append("Financial Statements (Insurance)")
    elif any(k in all_text for k in ["profit and loss", "p&l", "revenue from operations", "ebitda"]):
        parts.append("Statement of Profit & Loss")
    elif any(k in all_text for k in ["balance sheet", "total assets", "liabilities"]):
        parts.append("Balance Sheet")
    elif any(k in all_text for k in ["cash flow", "operating activities"]):
        parts.append("Cash Flow Statement")
    else:
        parts.append("Financial Statements")

    if year_headers:
        parts.append(f"— {year_headers[0]}")

    return " ".join(parts) if parts else "Financial Results"

def _detect_unit_label(raw_rows) -> str:
    combined = " ".join(r.get("item", "") for r in raw_rows[:30])
    if "₹" in combined or "crore" in combined.lower() or "inr" in combined.lower():
        return "₹ in Crores"
    if "lakhs" in combined.lower() or "lakh" in combined.lower():
        return "₹ in Lakhs"
    if "$" in combined or "usd" in combined.lower():
        return "USD in Millions"
    if "£" in combined or "gbp" in combined.lower():
        return "GBP in Millions"
    if "€" in combined or "eur" in combined.lower():
        return "EUR in Millions"
    return "Local Currency"

def _build_financial_sheet(wb, raw_rows, year_headers, N, unit_label, doc_title, doc_type="generic"):
    ws = wb.create_sheet("Financial Statements")
    _set_col_widths(ws, N)

    hdr_color, sub_color = _DOC_COLORS.get(doc_type, _DOC_COLORS["generic"])

    r = 1
    _title_bar(ws, r, N, f"  {doc_title}", hdr_color=hdr_color); r += 1
    _spacer(ws, r); r += 1
    _header_row(ws, r, year_headers, hdr_color=sub_color); r += 1
    _subheader_row(ws, r, N, f"({unit_label})", hdr_color=sub_color); r += 1
    _spacer(ws, r); r += 1
    ws.freeze_panes = f"B{r}"

    total_rev_row = None
    pat_rows = []

    for raw in raw_rows:
        item = clean_item_name(raw.get("item", ""))
        if not item:
            continue

        values_raw = raw.get("values", [])
        nums = []
        for v in values_raw[:N]:
            parsed = parse_number(str(v)) if v != "" and v is not None else None
            nums.append(parsed)
        while len(nums) < N:
            nums.append(None)

        has_values = any(isinstance(v, (int, float)) for v in nums)
        is_sec_flag = raw.get("is_section_header", False)
        row_type = _classify_row(item, has_values, is_sec_flag)
        indent = raw.get("indent", 0)
        is_pat_row = _is_pat(item)

        display = [v if v is not None else None for v in nums]

        if row_type == "SECTION":
            _section_header(ws, r, N, item)

        elif row_type == "TOTAL":
            bg = LT_BLUE if "profit" in item.lower() or "income" in item.lower() or "surplus" in item.lower() else PALE
            _total_row(ws, r, N, item, display, bg=bg, double_border=is_pat_row)

            lower = item.lower()
            if ("total revenue" in lower or "total income" in lower or "net premium" in lower) \
                    and total_rev_row is None:
                total_rev_row = r

            if is_pat_row:
                pat_rows.append(r)

            if total_rev_row and total_rev_row != r:
                formulas = [
                    f"=IFERROR({get_column_letter(c+2)}{r}/{get_column_letter(c+2)}{total_rev_row},\"\")"
                    for c in range(N)
                ]
                margin_label = None
                if "gross profit" in lower:
                    margin_label = "   Gross Margin %"
                elif "ebitda" in lower:
                    margin_label = "   EBITDA Margin %"
                elif "profit before tax" in lower or "pbt" in lower:
                    margin_label = "   PBT Margin %"
                elif is_pat_row:
                    margin_label = "   PAT Margin %"
                elif "net claims" in lower:
                    margin_label = "   Claims Ratio %"
                elif "underwriting" in lower and "surplus" in lower:
                    margin_label = "   Underwriting Margin %"

                if margin_label:
                    r += 1
                    _pct_row(ws, r, margin_label, formulas)

        else:
            _data_row(ws, r, item, display, indent=indent, bold=(indent == 0 and has_values))

        r += 1

    _spacer(ws, r, 4); r += 1
    note = ws.cell(row=r, column=1,
                   value="Note: All figures in above unit unless stated otherwise. "
                         "Figures in parentheses indicate negative values.")
    note.font = Font(italic=True, size=8, color="777777", name="Arial")

def _build_metrics_sheet(wb, raw_rows, year_headers, N, unit_label, doc_type="generic"):
    ws = wb.create_sheet("Key Metrics")
    _set_col_widths(ws, N)

    hdr_color, sub_color = _DOC_COLORS.get(doc_type, _DOC_COLORS["generic"])

    r = 1
    _title_bar(ws, r, N, "  KEY FINANCIAL METRICS", hdr_color=hdr_color); r += 1
    _spacer(ws, r); r += 1
    _header_row(ws, r, year_headers, hdr_color=sub_color); r += 1
    _subheader_row(ws, r, N, f"({unit_label})", hdr_color=sub_color); r += 1
    _spacer(ws, r); r += 1

    _KPI_MAP_PL = {
        "Total Revenue": ["total revenue from operations", "revenue from operations", "total income", "net premium earned"],
        "Other Income": ["other income"],
        "Total Income": ["total income"],
        "Total Expenses": ["total expenses"],
        "Finance Costs": ["finance costs", "finance cost"],
        "Depreciation & Amortisation": ["depreciation and amortisation", "depreciation"],
        "EBITDA": ["ebitda"],
        "Profit Before Tax": ["profit before tax", "profit before exceptional items and tax", "pbt"],
        "Tax Expense": ["total tax expense", "tax expense/(credit)", "income tax expense"],
        "PAT": ["profit for the period", "profit for the year", "profit/(loss) for the year"],
        "Total Comprehensive Income": ["total comprehensive income"],
    }
    _KPI_MAP_BS = {
        "Share Capital": ["paid-up", "share capital"],
        "Reserves & Surplus": ["reserves excluding revaluation", "reserves and surplus"],
        "Total Equity": ["total equity", "total shareholders equity"],
        "Non-Current Liabilities": ["total non-current liabilities"],
        "Current Liabilities": ["total current liabilities"],
        "Total Liabilities": ["total liabilities"],
        "Non-Current Assets": ["total non-current assets"],
        "Current Assets": ["total current assets"],
        "Total Assets": ["total assets"],
    }
    _KPI_MAP_CF = {
        "Net Cash — Operating": ["net cash from operating", "net cash generated from operating", "cash from operations"],
        "Net Cash — Investing": ["net cash from investing", "net cash used in investing"],
        "Net Cash — Financing": ["net cash from financing", "net cash used in financing"],
        "Net Change in Cash": ["net increase/decrease", "net change in cash"],
        "Opening Cash": ["opening cash", "cash at beginning"],
        "Closing Cash": ["closing cash", "cash at end", "cash and cash equivalents at the end"],
    }
    _KPI_MAP_INS = {
        "Gross Premium": ["gross premium", "total premium"],
        "Net Premium Earned": ["net premium earned"],
        "Net Claims Incurred": ["net claims incurred", "net claims"],
        "Underwriting Result": ["underwriting surplus", "underwriting deficit", "underwriting result"],
        "Investment Income": ["investment income", "income from investments"],
        "Profit Before Tax": ["profit before tax"],
        "PAT": ["profit after tax", "profit for the year"],
    }

    kpi_map = {
        "pl": _KPI_MAP_PL,
        "bs": _KPI_MAP_BS,
        "cf": _KPI_MAP_CF,
        "insurance": _KPI_MAP_INS,
    }.get(doc_type, _KPI_MAP_PL)

    item_lookup = {}
    for row in raw_rows:
        key = clean_item_name(row.get("item", "")).lower()
        item_lookup[key] = row

    def _find_row(candidates):
        for candidate in candidates:
            for key, row in item_lookup.items():
                if candidate in key:
                    return row
        return None

    def _get_vals(row):
        if not row:
            return [None] * N
        vals = []
        for v in row.get("values", [])[:N]:
            vals.append(parse_number(str(v)) if v != "" and v is not None else None)
        while len(vals) < N:
            vals.append(None)
        return vals[:N]

    _section_header(ws, r, N, "Key Line Items"); r += 1
    for label, candidates in kpi_map.items():
        found = _find_row(candidates)
        if found:
            _data_row(ws, r, label, _get_vals(found), bold=True); r += 1

    _spacer(ws, r); r += 1
    _section_header(ws, r, N, "Margin / Ratio Analysis"); r += 1

    fs_ws = wb["Financial Statements"]

    def _find_fs_row(keywords):
        for keyword in (keywords if isinstance(keywords, list) else [keywords]):
            for fs_row in fs_ws.iter_rows(min_row=1, max_row=fs_ws.max_row, min_col=1, max_col=1):
                cell_val = str(fs_row[0].value or "").lower()
                if keyword.lower() in cell_val:
                    return fs_row[0].row
        return None

    rev_fs_row = _find_fs_row(["total revenue", "net premium earned", "total income"])
    ebitda_fs_row = _find_fs_row(["ebitda"])
    pbt_fs_row = _find_fs_row(["profit before tax"])
    pat_fs_row = _find_fs_row(["profit for the period", "profit for the year"])
    claims_fs_row = _find_fs_row(["net claims incurred"])

    def _margin_formulas(num_row, den_row):
        if not num_row or not den_row:
            return [""] * N
        return [
            f"=IFERROR('Financial Statements'!{get_column_letter(c+2)}{num_row}/"
            f"'Financial Statements'!{get_column_letter(c+2)}{den_row},\"\")"
            for c in range(N)
        ]

    def _write_pct(label, num_row, den_row):
        nonlocal r
        fmls = _margin_formulas(num_row, den_row)
        if any(f for f in fmls):
            _data_row(ws, r, label, fmls)
            for c in range(N):
                ws.cell(row=r, column=c + 2).number_format = "0.0%"
                ws.cell(row=r, column=c + 2).font = Font(size=9, color=BLACK, name="Arial")
            r += 1

    if rev_fs_row:
        _write_pct("EBITDA Margin %", ebitda_fs_row, rev_fs_row)
        _write_pct("PBT Margin %", pbt_fs_row, rev_fs_row)
        _write_pct("PAT Margin %", pat_fs_row, rev_fs_row)
        _write_pct("Claims Ratio %", claims_fs_row, rev_fs_row)

    _spacer(ws, r, 4); r += 1
    note = ws.cell(row=r, column=1,
                   value="Margins are cross-sheet formulas linked to the Financial Statements sheet.")
    note.font = Font(italic=True, size=8, color="777777", name="Arial")

def _build_raw_sheet(wb, raw_rows, year_headers, N, unit_label):
    ws = wb.create_sheet("Raw Data")
    ws.column_dimensions["A"].width = 68
    for i in range(N):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16

    _title_bar(ws, 1, N, f"  RAW DATA  |  {unit_label}  |  Source: Uploaded PDF")

    ws.row_dimensions[2].height = 20
    c = ws.cell(row=2, column=1, value="Line Item")
    c.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill = _fill(MID_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for i, h in enumerate(year_headers):
        cc = ws.cell(row=2, column=i + 2, value=h)
        cc.font = Font(bold=True, size=9, color=WHITE, name="Arial")
        cc.fill = _fill(MID_BLUE)
        cc.alignment = Alignment(horizontal="right", vertical="center")

    rr = 3
    for row in raw_rows:
        item = clean_item_name(row.get("item", ""))
        if not item:
            continue

        is_sec = row.get("is_section_header", False)
        is_tot = _classify_row(item, True, False) == "TOTAL"
        vals = row.get("values", [])

        ws.row_dimensions[rr].height = 14
        c = ws.cell(row=rr, column=1, value=item)
        c.border = _bottom()

        if is_sec:
            c.font = Font(bold=True, size=9, color=NAVY, name="Arial")
            c.fill = _fill(SEC_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for ci in range(2, N + 2):
                cc = ws.cell(row=rr, column=ci)
                cc.fill = _fill(SEC_BG)
                cc.border = _bottom()
        elif is_tot:
            c.font = Font(bold=True, size=9, color=NAVY, name="Arial")
            c.fill = _fill(PALE)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            c.font = Font(size=9, color=DARK_GREY, name="Arial")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        for vi, v in enumerate(vals[:N]):
            if isinstance(v, str) and v.strip():
                v = parse_number(v) or v
            cell = ws.cell(row=rr, column=vi + 2, value=v)
            cell.border = _bottom()
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if is_sec:
                cell.fill = _fill(SEC_BG)
            elif is_tot:
                cell.fill = _fill(PALE)

            if isinstance(v, (int, float)):
                if abs(v) < 500 and "eps" in item.lower():
                    cell.number_format = '0.00;(0.00);"-"'
                else:
                    cell.number_format = '#,##0.00;(#,##0.00);"-"'
                cell.font = Font(
                    bold=is_tot, size=9,
                    color=NAVY if is_tot else BLUE_NUM,
                    name="Arial"
                )
            else:
                cell.font = Font(size=9, color=DARK_GREY, name="Arial")

        rr += 1

    ws.row_dimensions[rr + 1].height = 4
    note = ws.cell(row=rr + 2, column=1,
                   value="Source: Extracted from uploaded PDF. "
                         "Blue numbers = hardcoded inputs. Bold rows = totals/subtotals.")
    note.font = Font(italic=True, size=8, color="777777", name="Arial")

def create_excel(processed_data: dict) -> BytesIO:
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    year_headers = processed_data.get("year_headers") or []
    raw_rows = processed_data.get("raw_rows") or []

    for row in raw_rows:
        normed = []
        for v in row.get("values", []):
            p = parse_number(str(v)) if v is not None and v != "" else None
            normed.append(p)
        row["values"] = normed

    if not year_headers:
        year_headers = ["Period 1", "Period 2", "Period 3", "Period 4", "Period 5"]

    N = len(year_headers)
    unit_label = _detect_unit_label(raw_rows)
    doc_title = _detect_doc_title(raw_rows, year_headers)
    doc_type = _detect_doc_type(raw_rows)

    _build_financial_sheet(wb, raw_rows, year_headers, N, unit_label, doc_title, doc_type)
    _build_metrics_sheet(wb, raw_rows, year_headers, N, unit_label, doc_type)
    _build_raw_sheet(wb, raw_rows, year_headers, N, unit_label)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def _create_raw_sheet(wb, raw_rows, headers):
    N = len(headers)
    _build_raw_sheet(wb, raw_rows, headers, N, _detect_unit_label(raw_rows))

def _create_pl_sheet(wb, raw_rows, year_headers, N):
    unit = _detect_unit_label(raw_rows)
    title = _detect_doc_title(raw_rows, year_headers)
    dtype = _detect_doc_type(raw_rows)
    _build_financial_sheet(wb, raw_rows, year_headers, N, unit, title, dtype)